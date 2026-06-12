from zoneinfo import ZoneInfo
import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse, parse_qs, urlencode, urlunparse

import requests
from openpyxl import load_workbook


EXCEL_SHARE_URL = os.environ.get("EXCEL_SHARE_URL", "").strip()
EXCEL_FILE = Path("latest-vm-tips.xlsm")
OUTPUT_FILE = Path("resultat.json")

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/120 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

def add_download_param(url):
    parsed = urlparse(url)
    query = parse_qs(parsed.query)
    query["download"] = ["1"]
    return urlunparse(parsed._replace(query=urlencode(query, doseq=True)))

def download_like_browser():
    if not EXCEL_SHARE_URL:
        raise RuntimeError("EXCEL_SHARE_URL saknas")

    session = requests.Session()
    session.headers.update(HEADERS)

    candidates = [EXCEL_SHARE_URL, add_download_param(EXCEL_SHARE_URL)]
    last_error = None

    for url in candidates:
        try:
            print(f"Testar: {url}")
            r = session.get(url, allow_redirects=True, timeout=90)
            print(f"Status: {r.status_code}")
            print(f"Slutlig URL: {r.url}")
            print(f"Content-Type: {r.headers.get('content-type')}")
            print(f"Storlek: {len(r.content)} bytes")

            if r.content.startswith(b"PK") and len(r.content) > 50000:
                EXCEL_FILE.write_bytes(r.content)
                print("Excel-fil hämtad direkt.")
                return

            html = r.text
            possible_urls = set(re.findall(r'https?:\\/\\/[^"\\]+', html))
            possible_urls.update(re.findall(r'https?://[^"\']+', html))

            cleaned = []
            for u in possible_urls:
                u = u.replace("\\/", "/").replace("\\u0026", "&")
                if any(token in u.lower() for token in ["download", "contents", "download.aspx", ".xlsm", ".xlsx"]):
                    cleaned.append(u)

            print(f"Hittade {len(cleaned)} möjliga fil-URL:er i HTML.")

            for direct in cleaned[:30]:
                print(f"Testar hittad URL: {direct[:180]}")
                rr = session.get(direct, allow_redirects=True, timeout=90)
                print(f"  Status: {rr.status_code}, typ: {rr.headers.get('content-type')}, storlek: {len(rr.content)}")
                if rr.content.startswith(b"PK") and len(rr.content) > 50000:
                    EXCEL_FILE.write_bytes(rr.content)
                    print("Excel-fil hämtad via hittad URL.")
                    return

            last_error = f"Ingen Excel-fil hittades via {url}"

        except Exception as exc:
            last_error = str(exc)
            print(f"Misslyckades: {exc}")

    raise RuntimeError(last_error or "Kunde inte hämta Excel-filen")

def v(ws, row, col):
    return ws.cell(row=row, column=col).value

def read_leaderboard(ws):
    leaderboard = []
    for row in range(4, 90):
        position = v(ws, row, 208)
        name = v(ws, row, 210)
        points = v(ws, row, 211)
        if name and points is not None:
            try:
                position = int(position)
            except Exception:
                position = len(leaderboard) + 1
            try:
                points = int(points)
            except Exception:
                pass
            leaderboard.append({"position": position, "name": str(name).strip(), "points": points})
    if not leaderboard:
        raise RuntimeError("Hittade ingen topplista.")
    return leaderboard

def read_matches(ws):
    matches = []
    current_section = ""
    for row in range(2, 170):
        b = v(ws, row, 2)
        if isinstance(b, str):
            text = b.strip()
            if text and text not in ("Match", "Gruppspel / slutspel", "Fyll i resultat"):
                current_section = text

        match_no = v(ws, row, 2)
        date_val = v(ws, row, 3)
        home = v(ws, row, 4)
        away = v(ws, row, 6)
        home_goals = v(ws, row, 7)
        away_goals = v(ws, row, 9)

        if isinstance(match_no, (int, float)) and home and away:
            date = date_val.strftime("%Y-%m-%d") if hasattr(date_val, "strftime") else (str(date_val) if date_val else "")
            played = home_goals is not None and away_goals is not None
            if played:
                try:
                    hg, ag = int(home_goals), int(away_goals)
                    score = f"{hg}–{ag}"
                except Exception:
                    hg, ag = home_goals, away_goals
                    score = f"{home_goals}–{away_goals}"
            else:
                hg, ag, score = None, None, "–"

            matches.append({
                "section": current_section,
                "match": int(match_no),
                "date": date,
                "home": str(home).strip(),
                "away": str(away).strip(),
                "homeGoals": hg,
                "awayGoals": ag,
                "score": score,
                "status": "spelad" if played else "kommande"
            })
    return matches

def main():
    download_like_browser()
    workbook = load_workbook(EXCEL_FILE, data_only=True, keep_vba=True)
    if "Resultat & tabell" not in workbook.sheetnames:
        raise RuntimeError(f"Fliken 'Resultat & tabell' saknas. Flikar: {workbook.sheetnames}")
    ws = workbook["Resultat & tabell"]

    data = {
        "updated": datetime.now(ZoneInfo("Europe/Stockholm")).strftime("%Y-%m-%d %H:%M"),
        "leaderboard": read_leaderboard(ws),
        "matches": read_matches(ws),
    }
    OUTPUT_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Skapade {OUTPUT_FILE}")
    print(f"Deltagare: {len(data['leaderboard'])}")
    print(f"Matcher: {len(data['matches'])}")
    print(f"Spelade matcher: {sum(1 for m in data['matches'] if m['status'] == 'spelad')}")

if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"FEL: {exc}", file=sys.stderr)
        sys.exit(1)
